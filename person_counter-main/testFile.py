import cv2
import openpyxl
from os import walk
import re
from openpyxl import load_workbook

#wb_obj = openpyxl.load_workbook("performance.xlsx")





# file1 = open('./file.txt', 'r');
# lines = file1.readlines()

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    '''
    alist.sort(key=natural_keys) sorts in human order
    http://nedbatchelder.com/blog/200712/human_sorting.html
    (See Toothy's implementation in the comments)
    '''
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]




# This is the path where all the files are stored.
folder_path = '/Users/softedel/Desktop/person_counter-main/videos'
f=[]
# Open one of the files,
com=1
for (dirpath, dirnames, filenames) in walk(folder_path):

    f.append(filenames)
    f[0].sort(key=natural_keys)
    workbook = load_workbook(filename="demo.xlsx")
    sheet = workbook.active

    # Write what you want into a specific cell
  #   for fil in f[0]:
  #       column = "e" + str(com)
  #       sheet[column] = f[0][com]
  #       print(f[0][com])
  #       com = com + 1
  #       # Save the spreadsheet
  #       workbook.save(filename="demo.xlsx")
  # #  print(f[0])
  #   break




#
for line in f[0]:
     videoName= line[0:15]
     cap = cv2.VideoCapture("videos/"+videoName)
     while (True):
         ret, frame = cap.read()
         # show the output frame
         try:
             cv2.imshow('result', frame)
             key = cv2.waitKey(1) & 0xFF

              # if the `q` key was pressed, break from the loop
             if key == ord("q"):
                 break

         except:
             break
             print("Here")






#          key = cv2.waitKey(1) & 0xFF
#
#         # # if the `q` key was pressed, break from the loop
#          if key == ord("q"):
#              break
#
# print("Line{}: {}".format(count, line.strip()))
